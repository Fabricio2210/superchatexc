import openpyxl
def excel(file, name, subject):
    workbook = openpyxl.Workbook()
    sheetSubscription = workbook.active
    sheetSubscription.title = "subscription"
    sheetSuperchats = workbook.create_sheet("superchats")
    sheetGiftedMembership = workbook.create_sheet("Gifted_Membership")
    column_widths = {
        'A': 50,
        'B': 100,
        'C': 50,
        'D': 50,
        'E': 12,
        'F': 100,
    }

    for column, width in column_widths.items():
        sheetSubscription.column_dimensions[column].width = width 
        sheetSuperchats.column_dimensions[column].width = width
        sheetGiftedMembership.column_dimensions[column].width = width

    sheetSubscription.append(['ID', 'Profile Picture', 'Name', 'Message'])
    sheetGiftedMembership.append(['ID', 'Profile Picture', 'Name', 'Message'])
    sheetSuperchats.append(['ID', 'Profile Picture', 'Name', 'Money', 'Currency', 'Message'])
    for data in file:
        if not data.get('money') and not data.get('isGifted'):
            sheetSubscription.append([data.get('id'), data.get('profilePicture'), data.get('name'), data.get('message')])
        elif data.get('isGifted'):
            sheetGiftedMembership.append([data.get('id'), data.get('profilePicture'), data.get('name'), data.get('message')])
        elif data.get('money') and not data.get('isGifted'):
            sheetSuperchats.append([data.get("id"), data.get("profilePicture"), data.get("name"), data.get("money"), data.get("currency"), data.get("message")])
    parsedName = name.split(".")
    workbook.save(f'./{subject}/excel/{parsedName[0]}.xlsx')
